from copy import copy
from datetime import datetime
from openpyxl import load_workbook

from config import (
    MODELO_EXCEL,
    PASTA_PLANILHAS,
    FATURADO_POR,
    LINHA_INICIAL
)


class ExcelWriter:

    def __init__(self):

        self.wb = None
        self.ws = None

    # ----------------------------------------------------

    def carregar_modelo(self):

        self.wb = load_workbook(MODELO_EXCEL)

        self.ws = self.wb.active

    # ----------------------------------------------------

    def copiar_linha(self, origem, destino):

        """
        Copia completamente uma linha do modelo.
        """

        self.ws.row_dimensions[destino].height = (
            self.ws.row_dimensions[origem].height
        )

        for coluna in range(1, 13):

            origem_cell = self.ws.cell(origem, coluna)
            destino_cell = self.ws.cell(destino, coluna)

            if origem_cell.has_style:
                destino_cell._style = copy(origem_cell._style)

            destino_cell.font = copy(origem_cell.font)
            destino_cell.fill = copy(origem_cell.fill)
            destino_cell.border = copy(origem_cell.border)
            destino_cell.alignment = copy(origem_cell.alignment)
            destino_cell.number_format = origem_cell.number_format
            destino_cell.protection = copy(origem_cell.protection)

            # Copia também fórmulas existentes
            destino_cell.value = origem_cell.value

    # ----------------------------------------------------

    def preparar_planilha(self, quantidade_produtos):

        """
        Cria somente as linhas necessárias.
        """

        if quantidade_produtos <= 1:
            return

        for indice in range(quantidade_produtos - 1):

            linha = LINHA_INICIAL + indice + 1

            self.ws.insert_rows(linha)

            self.copiar_linha(LINHA_INICIAL, linha)

    # ----------------------------------------------------

    def escrever_produtos(self, pedido):

        """
        Preenche somente as colunas referentes
        aos produtos.
        """

        linha = LINHA_INICIAL

        for produto in pedido.produtos:

            if linha == LINHA_INICIAL:
                self.ws[f"A{linha}"] = pedido.nome_cliente

            self.ws[f"B{linha}"] = produto.descricao

            self.ws[f"C{linha}"] = produto.cx_kg

            self.ws[f"D{linha}"] = produto.volume

            self.ws[f"F{linha}"] = produto.venda

            linha += 1

    # ----------------------------------------------------

    def escrever_dados_pedido(self, pedido):

        ultima_linha = LINHA_INICIAL + len(pedido.produtos) - 1

        # Mesclagens
        if ultima_linha > LINHA_INICIAL:

            self.ws.merge_cells(
                start_row=LINHA_INICIAL,
                start_column=8,
                end_row=ultima_linha,
                end_column=8
            )

            self.ws.merge_cells(
                start_row=LINHA_INICIAL,
                start_column=9,
                end_row=ultima_linha,
                end_column=9
            )

            self.ws.merge_cells(
                start_row=LINHA_INICIAL,
                start_column=10,
                end_row=ultima_linha,
                end_column=10
            )

            self.ws.merge_cells(
                start_row=LINHA_INICIAL,
                start_column=11,
                end_row=ultima_linha,
                end_column=11
            )

        self.ws[f"H{LINHA_INICIAL}"] = datetime.now().strftime("%d/%m/%Y")

        self.ws[f"I{LINHA_INICIAL}"] = pedido.prazo_pgto

        self.ws[f"J{LINHA_INICIAL}"] = pedido.entrega

        self.ws[f"K{LINHA_INICIAL}"] = FATURADO_POR

        # Centraliza os campos mesclados
        for coluna in ["H", "I", "J", "K"]:

            self.ws[f"{coluna}{LINHA_INICIAL}"].alignment = copy(
                self.ws[f"{coluna}{LINHA_INICIAL}"].alignment
            )

    # ----------------------------------------------------

    def finalizar(self, pedido):

        ultima_linha = LINHA_INICIAL + len(pedido.produtos) - 1

        # Remove linhas abaixo do último produto
        max_linhas = self.ws.max_row

        if ultima_linha < max_linhas:

            self.ws.delete_rows(
                ultima_linha + 1,
                max_linhas - ultima_linha
            )

    # ----------------------------------------------------

    def salvar(self, pedido):

        data = datetime.now().strftime("%d-%m-%y")

        nome = f"PEDIDO {pedido.nome_cliente} FRIGORIFICO {data}.xlsx"

        caracteres = '\\/:*?"<>|'

        for c in caracteres:
            nome = nome.replace(c, "")

        PASTA_PLANILHAS.mkdir(exist_ok=True)

        destino = PASTA_PLANILHAS / nome

        self.wb.save(destino)

        self.wb.close()

        return destino

    # ----------------------------------------------------

    def gerar(self, pedido):

        self.carregar_modelo()

        self.preparar_planilha(len(pedido.produtos))

        self.escrever_produtos(pedido)

        self.escrever_dados_pedido(pedido)

        self.finalizar(pedido)

        return self.salvar(pedido)