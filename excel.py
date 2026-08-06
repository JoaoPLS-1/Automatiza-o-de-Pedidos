from copy import copy
from datetime import datetime
from openpyxl import load_workbook

from config import (
    MODELO_EXCEL,
    PASTA_PLANILHAS,
)


class ExcelWriter:

    def __init__(self):
        self.wb = None
        self.ws = None

    def carregar_modelo(self):
        self.wb = load_workbook(MODELO_EXCEL)
        self.ws = self.wb.active

    def copiar_linha(self, origem, destino):

        for col in range(1, 13):

            origem_cell = self.ws.cell(origem, col)
            destino_cell = self.ws.cell(destino, col)

            destino_cell._style = copy(origem_cell._style)
            destino_cell.font = copy(origem_cell.font)
            destino_cell.fill = copy(origem_cell.fill)
            destino_cell.border = copy(origem_cell.border)
            destino_cell.alignment = copy(origem_cell.alignment)
            destino_cell.number_format = origem_cell.number_format
            destino_cell.protection = copy(origem_cell.protection)

    def gerar(self, pedido):

        self.carregar_modelo()

        data = datetime.now().strftime("%d-%m-%y")

        nome = f"PEDIDO {pedido.nome_cliente} FRIGORIFICO {data}.xlsx"

        nome = (
            nome.replace("/", "")
                .replace("\\", "")
                .replace(":", "")
                .replace("*", "")
                .replace("?", "")
                .replace('"', "")
                .replace("<", "")
                .replace(">", "")
                .replace("|", "")
        )

        PASTA_PLANILHAS.mkdir(exist_ok=True)

        destino = PASTA_PLANILHAS / nome

        self.wb.save(destino)

        self.wb.close()

        return destino